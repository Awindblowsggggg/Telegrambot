import json
import os
import re
from datetime import datetime
from telegram import Update, ReplyKeyboardMarkup, ReplyKeyboardRemove
from telegram.ext import (
    Application,
    CommandHandler,
    MessageHandler,
    ConversationHandler,
    CallbackContext,
    filters
)
from openpyxl import load_workbook
from openpyxl import Workbook

# Definición de los estados del formulario
(
    ID_CARRO, CONDICION, TIPO, FECHA, HORA, 
    AM_PM, IMPORTE, CONDUCTOR1, CONDUCTOR2, 
    OBSERVACION, TONELADAS
) = range(11)

# Archivos de datos
FORMULARIOS_FILE = "formularios.json"
EXCEL_FILE = "formulario.xlsx"

# --- Funciones auxiliares ---

def cargar_formularios():
    """Carga los formularios desde el archivo JSON"""
    try:
        if os.path.exists(FORMULARIOS_FILE):
            with open(FORMULARIOS_FILE, 'r', encoding='utf-8') as file:
                return json.load(file)
        return {}
    except (json.JSONDecodeError, FileNotFoundError):
        return {}

def guardar_formularios(formularios):
    """Guarda los formularios en el archivo JSON"""
    with open(FORMULARIOS_FILE, 'w', encoding='utf-8') as file:
        json.dump(formularios, file, ensure_ascii=False, indent=2)

def obtener_ultimo_registro(id_carro):
    """Obtiene el último registro para un ID de carro específico"""
    formularios = cargar_formularios()
    ultimo_registro = None
    
    for form in formularios.values():
        if form.get('id_carro') == id_carro:
            if not ultimo_registro:
                ultimo_registro = form
            else:
                # Comparar fechas y horas para determinar cuál es más reciente
                try:
                    fecha_actual = datetime.strptime(form['fecha'], '%d/%m/%Y')
                    fecha_ultima = datetime.strptime(ultimo_registro['fecha'], '%d/%m/%Y')
                    
                    if fecha_actual > fecha_ultima:
                        ultimo_registro = form
                    elif fecha_actual == fecha_ultima:
                        # Si es el mismo día, comparar horas
                        hora_actual = form['hora_completa']
                        hora_ultima = ultimo_registro['hora_completa']
                        if hora_actual > hora_ultima:
                            ultimo_registro = form
                except ValueError:
                    continue
    
    return ultimo_registro

def guardar_en_excel(formulario):
    """Guarda los datos en el archivo Excel"""
    try:
        if os.path.exists(EXCEL_FILE):
            wb = load_workbook(EXCEL_FILE)
        else:
            wb = Workbook()
        
        ws = wb.active
        
        # Buscar primera fila vacía
        fila = 1
        while ws[f'A{fila}'].value is not None:
            fila += 1
            
        # Mapeo de campos a columnas
        campos = [
            'id_carro', 'condicion', 'tipo', 'fecha', 
            'hora_completa', 'importe', 'conductor1', 
            'conductor2', 'observacion', 'toneladas'
        ]
        
        for col, campo in enumerate(campos, start=1):
            ws.cell(row=fila, column=col, value=formulario.get(campo, 'No especificado'))
        
        wb.save(EXCEL_FILE)
    except Exception as e:
        print(f"Error al guardar en Excel: {e}")

# --- Handlers del formulario ---

async def start(update: Update, context: CallbackContext) -> int:
    """Inicia el formulario"""
    reply_keyboard = [
        ['4227', '4242', '4243'],
        ['4245', '4254', '5441'],
        ['5524', '9061', '9063'],
        ['VW1', 'VW2', 'VW3']
    ]
    await update.message.reply_text(
        "¡Hola! Vamos a comenzar el formulario 🧾\n\n"
        "1️⃣. Selecciona el ID del Carro 🚛:",
        reply_markup=ReplyKeyboardMarkup(
            reply_keyboard, 
            one_time_keyboard=True,
            input_field_placeholder="ID del carro"
        )
    )
    return ID_CARRO

async def id_carro(update: Update, context: CallbackContext) -> int:
    """Maneja la selección del ID del carro"""
    context.user_data['id_carro'] = update.message.text
    reply_keyboard = [
        ['Listo', 'MC', 'MP'],
        ['PH', 'Op', 'T.Inact'],
        ['ESP']
    ]
    await update.message.reply_text(
        "2️⃣. Selecciona la condición ♻️:",
        reply_markup=ReplyKeyboardMarkup(
            reply_keyboard, 
            one_time_keyboard=True,
            input_field_placeholder="Condición del carro"
        )
    )
    return CONDICION

async def condicion(update: Update, context: CallbackContext) -> int:
    """Maneja la condición del carro"""
    context.user_data['condicion'] = update.message.text
    reply_keyboard = [['Inicio', 'Fin']]
    await update.message.reply_text(
        "3️⃣. Elige el tipo (✅Inicio/❌Fin):",
        reply_markup=ReplyKeyboardMarkup(
            reply_keyboard,
            one_time_keyboard=True,
            input_field_placeholder="Tipo de registro"
        )
    )
    return TIPO

async def tipo(update: Update, context: CallbackContext) -> int:
    """Valida y maneja el tipo de registro (Inicio/Fin)"""
    id_carro = context.user_data['id_carro']
    tipo_seleccionado = update.message.text
    ultimo_registro = obtener_ultimo_registro(id_carro)
    
    # Validación de secuencia Inicio-Fin
    if tipo_seleccionado == "Inicio":
        if ultimo_registro and ultimo_registro['tipo'] == "Inicio":
            await update.message.reply_text(
                f"⛔ ERROR: El carro {id_carro} ya tiene un Inicio pendiente.\n"
                f"📅 Fecha: {ultimo_registro['fecha']}\n"
                f"⏰ Hora: {ultimo_registro['hora_completa']}\n\n"
                "Debes registrar un Fin antes de iniciar otro turno.",
                reply_markup=ReplyKeyboardRemove()
            )
            return ConversationHandler.END
            
    elif tipo_seleccionado == "Fin":
        if not ultimo_registro or ultimo_registro['tipo'] == "Fin":
            msg = (f"⛔ ERROR: No hay un Inicio pendiente para {id_carro}." 
                  if not ultimo_registro 
                  else f"⛔ ERROR: El último registro para {id_carro} ya es un Fin.")
            
            await update.message.reply_text(
                msg + "\n\nDebes registrar un Inicio primero.",
                reply_markup=ReplyKeyboardRemove()
            )
            return ConversationHandler.END

    context.user_data['tipo'] = tipo_seleccionado
    await update.message.reply_text(
        "4️⃣. Introduce la fecha (día/mes/año) 📆:\n"
        "Ejemplo: 16/03/2025",
        reply_markup=ReplyKeyboardRemove()
    )
    return FECHA

async def fecha(update: Update, context: CallbackContext) -> int:
    """Valida y maneja la fecha"""
    fecha = update.message.text
    try:
        datetime.strptime(fecha, '%d/%m/%Y')
        context.user_data['fecha'] = fecha
        await update.message.reply_text(
            "5️⃣. Introduce la hora (hora:minutos) ⏰:\n"
            "Ejemplo: 02:30"
        )
        return HORA
    except ValueError:
        await update.message.reply_text(
            "Formato incorrecto. Usa día/mes/año\n"
            "Ejemplo: 16/03/2025"
        )
        return FECHA

async def hora(update: Update, context: CallbackContext) -> int:
    """Valida y maneja la hora"""
    hora = update.message.text
    if re.match(r'^\d{1,2}:\d{2}$', hora):
        context.user_data['hora'] = hora
        reply_keyboard = [['AM', 'PM']]
        await update.message.reply_text(
            "Selecciona AM o PM:",
            reply_markup=ReplyKeyboardMarkup(
                reply_keyboard,
                one_time_keyboard=True
            )
        )
        return AM_PM
    else:
        await update.message.reply_text(
            "Formato incorrecto. Usa hora:minutos\n"
            "Ejemplo: 02:30"
        )
        return HORA

async def am_pm(update: Update, context: CallbackContext) -> int:
    """Maneja AM/PM y construye hora completa"""
    context.user_data['am_pm'] = update.message.text
    context.user_data['hora_completa'] = (
        f"{context.user_data['hora']} {context.user_data['am_pm']}"
    )
    await update.message.reply_text(
        "6️⃣. Introduce el importe 💰 (solo números):"
    )
    return IMPORTE

async def importe(update: Update, context: CallbackContext) -> int:
    """Valida y maneja el importe"""
    importe = update.message.text
    if importe.isdigit():
        context.user_data['importe'] = int(importe)
        reply_keyboard = [
            ['Raidel Castel Neyra', 'Serguei Lago López'],
            ['Oscar Cruz Figueredo', 'Leonides Ruiz Núñez'],
            ['Carlos Fonseca Tamayo', 'Diosdado Quezada Aguilera'],
            ['Lienmi Almarales Oña', 'Reynaldo Real Almarales']
        ]
        await update.message.reply_text(
            "7️⃣. Selecciona al Conductor 1 🚗:",
            reply_markup=ReplyKeyboardMarkup(
                reply_keyboard,
                one_time_keyboard=True
            )
        )
        return CONDUCTOR1
    else:
        await update.message.reply_text("Debe ser un número. Intenta de nuevo:")
        return IMPORTE

async def conductor1(update: Update, context: CallbackContext) -> int:
    """Maneja conductor 1"""
    context.user_data['conductor1'] = update.message.text
    reply_keyboard = [
        ['Raidel Castel Neyra', 'Serguei Lago López'],
        ['Oscar Cruz Figueredo', 'Leonides Ruiz Núñez'],
        ['Carlos Fonseca Tamayo', 'Diosdado Quezada Aguilera'],
        ['Lienmi Almarales Oña', 'Reynaldo Real Almarales']
    ]
    await update.message.reply_text(
        "8️⃣. Selecciona al Conductor 2 🚗:",
        reply_markup=ReplyKeyboardMarkup(
            reply_keyboard,
            one_time_keyboard=True
        )
    )
    return CONDUCTOR2

async def conductor2(update: Update, context: CallbackContext) -> int:
    """Maneja conductor 2"""
    context.user_data['conductor2'] = update.message.text
    await update.message.reply_text(
        "9️⃣. Escribe cualquier observación 📝:"
    )
    return OBSERVACION

async def observacion(update: Update, context: CallbackContext) -> int:
    """Maneja observaciones"""
    context.user_data['observacion'] = update.message.text
    await update.message.reply_text(
        "🔟. Introduce las toneladas ⚖️ (1-100):"
    )
    return TONELADAS

async def toneladas(update: Update, context: CallbackContext) -> int:
    """Valida y maneja toneladas, finaliza el formulario"""
    toneladas = update.message.text
    if toneladas.isdigit() and 1 <= int(toneladas) <= 100:
        context.user_data['toneladas'] = int(toneladas)
        
        # Construir resumen detallado
        resumen = (
            "📋 FORMULARIO COMPLETADO\n\n"
            f"🚛 ID Carro: {context.user_data['id_carro']}\n"
            f"♻️ Condición: {context.user_data['condicion']}\n"
            f"🔄 Tipo: {context.user_data['tipo']}\n"
            f"📅 Fecha: {context.user_data['fecha']}\n"
            f"⏰ Hora: {context.user_data['hora_completa']}\n"
            f"💰 Importe: {context.user_data['importe']}\n"
            f"👤 Conductor 1: {context.user_data['conductor1']}\n"
            f"👤 Conductor 2: {context.user_data['conductor2']}\n"
            f"📝 Observación: {context.user_data['observacion']}\n"
            f"⚖️ Toneladas: {context.user_data['toneladas']}\n\n"
            f"📌 Registrado por: {update.effective_user.full_name}"
        )
        
        # Guardar datos
        formularios = cargar_formularios()
        formulario_id = (
            f"{context.user_data['id_carro']}_"
            f"{context.user_data['fecha']}_"
            f"{context.user_data['hora_completa']}"
        )
        formularios[formulario_id] = context.user_data
        guardar_formularios(formularios)
        guardar_en_excel(context.user_data)
        
        # Enviar resumen
        await update.message.reply_text(
            resumen,
            reply_markup=ReplyKeyboardRemove()
        )
        
        # Enviar al grupo (reemplaza con tu ID de grupo real)
        grupo_id = -1002459560918
        try:
            await context.bot.send_message(
                chat_id=grupo_id,
                text=resumen
            )
        except Exception as e:
            print(f"Error al enviar al grupo: {e}")
        
        return ConversationHandler.END
    else:
        await update.message.reply_text(
            "Debe ser un número entre 1 y 100. Intenta de nuevo:"
        )
        return TONELADAS

async def cancel(update: Update, context: CallbackContext) -> int:
    """Cancela el formulario"""
    await update.message.reply_text(
        "Formulario cancelado.",
        reply_markup=ReplyKeyboardRemove()
    )
    return ConversationHandler.END

async def ver_estados(update: Update, context: CallbackContext):
    """Muestra los estados actuales de todos los carros"""
    formularios = cargar_formularios()
    estados = {}
    
    # Encontrar el último estado de cada carro
    for form in formularios.values():
        id_carro = form['id_carro']
        if id_carro not in estados:
            estados[id_carro] = form
        else:
            # Comparar fechas para determinar cuál es más reciente
            try:
                fecha_actual = datetime.strptime(form['fecha'], '%d/%m/%Y')
                fecha_guardada = datetime.strptime(estados[id_carro]['fecha'], '%d/%m/%Y')
                
                if fecha_actual > fecha_guardada:
                    estados[id_carro] = form
                elif fecha_actual == fecha_guardada:
                    # Si es el mismo día, comparar horas
                    if form['hora_completa'] > estados[id_carro]['hora_completa']:
                        estados[id_carro] = form
            except ValueError:
                continue
    
    if not estados:
        await update.message.reply_text("No hay registros activos.")
        return
    
    # Construir mensaje organizado
    mensaje = "🚚 ESTADO ACTUAL DE LOS CARROS\n\n"
    
    # Separar en Inicios pendientes y Finalizados
    inicios_pendientes = []
    finalizados = []
    
    for id_carro, form in estados.items():
        if form['tipo'] == "Inicio":
            inicios_pendientes.append((id_carro, form))
        else:
            finalizados.append((id_carro, form))
    
    # Mostrar primero los Inicios pendientes
    if inicios_pendientes:
        mensaje += "⏳ INICIOS PENDIENTES:\n"
        for id_carro, form in inicios_pendientes:
            mensaje += (
                f"  🚛 {id_carro} - {form['condicion']}\n"
                f"    📅 {form['fecha']} ⏰ {form['hora_completa']}\n"
                f"    👤 {form['conductor1']}/{form['conductor2']}\n\n"
            )
    
    # Luego los Finalizados
    if finalizados:
        mensaje += "✅ FINALIZADOS:\n"
        for id_carro, form in finalizados:
            mensaje += (
                f"  🚛 {id_carro} - {form['condicion']}\n"
                f"    📅 {form['fecha']} ⏰ {form['hora_completa']}\n\n"
            )
    
    await update.message.reply_text(mensaje)

def main():
    """Configura y ejecuta el bot"""
    # Primero construimos la aplicación
    application = Application.builder().token("7915698639:AAGvT0Ouwth1AYAN6oC98u_fJa8O1lL2h7w").build()
    
    # Manejador de conversación para el formulario
    conv_handler = ConversationHandler(
        entry_points=[CommandHandler('start', start)],
        states={
            ID_CARRO: [MessageHandler(filters.TEXT & ~filters.COMMAND, id_carro)],
            CONDICION: [MessageHandler(filters.TEXT & ~filters.COMMAND, condicion)],
            TIPO: [MessageHandler(filters.TEXT & ~filters.COMMAND, tipo)],
            FECHA: [MessageHandler(filters.TEXT & ~filters.COMMAND, fecha)],
            HORA: [MessageHandler(filters.TEXT & ~filters.COMMAND, hora)],
            AM_PM: [MessageHandler(filters.TEXT & ~filters.COMMAND, am_pm)],
            IMPORTE: [MessageHandler(filters.TEXT & ~filters.COMMAND, importe)],
            CONDUCTOR1: [MessageHandler(filters.TEXT & ~filters.COMMAND, conductor1)],
            CONDUCTOR2: [MessageHandler(filters.TEXT & ~filters.COMMAND, conductor2)],
            OBSERVACION: [MessageHandler(filters.TEXT & ~filters.COMMAND, observacion)],
            TONELADAS: [MessageHandler(filters.TEXT & ~filters.COMMAND, toneladas)],
        },
        fallbacks=[CommandHandler('cancel', cancel)],
    )
    
    # Añadir los handlers a la aplicación
    application.add_handler(CommandHandler('estados', ver_estados))
    application.add_handler(conv_handler)
    
    # Ejecutar el bot
    print("Bot iniciado... Esperando mensajes")
    application.run_polling()
if __name__ == '__main__':
    # Verificar si existen los archivos necesarios
    if not os.path.exists(FORMULARIOS_FILE):
        with open(FORMULARIOS_FILE, 'w', encoding='utf-8') as f:
            json.dump({}, f)
    
    if not os.path.exists(EXCEL_FILE):
        wb = Workbook()
        ws = wb.active
        # Escribir los encabezados
        encabezados = [
            'ID Carro', 'Condición', 'Tipo', 'Fecha', 
            'Hora', 'Importe', 'Conductor 1', 
            'Conductor 2', 'Observación', 'Toneladas'
        ]
        for col, encabezado in enumerate(encabezados, start=1):
            ws.cell(row=1, column=col, value=encabezado)
        wb.save(EXCEL_FILE)
    
    main()
